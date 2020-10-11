import random
import openpyxl as xl
n = 10000
wb = xl.load_workbook('Demo.xlsx')
wb1 = xl.load_workbook('Demo.xlsx')
sheet = wb['Sheet']
sheet1 = wb1['Sheet']
cell1 = sheet.cell(1, 1)
cell1.value = "Aadhar-No"

celli = sheet1.cell(1, 1)
celli.value = "Aadhar-No"

cell2 = sheet.cell(1, 2)
cell2.value = "Ration-card"

cell3 = sheet.cell(1, 3)
cell3.value = "Age"

cell4 = sheet.cell(1, 4)
cell4.value = "Phone-NUmber"
for i in range(2, n):
    aa = "Ap"+str(random.randint(111111111, 999999999))
    cell = sheet.cell(i, 1)
    cell.value = aa
    celli = sheet1.cell(i, 1)
    celli.value = aa
    bb = "RA" + str(random.randint(111111111, 999999999))
    cell = sheet.cell(i, 2)
    cell.value = bb
    cc = random.randint(18, 100)
    cell = sheet.cell(i, 3)
    cell.value = cc
    dd = random.randint(7000000000, 9999999999)
    cell = sheet.cell(i, 4)
    cell.value = dd
wb.save("source.xlsx")
wb1.save('target.xlsx')
