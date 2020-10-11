import openpyxl as xl
import pandas as pd

# Loading Doccuments
df = pd.read_excel('source.xlsx')
target = xl.load_workbook('target.xlsx')
target_sheet = target['Sheet']
length = target_sheet.max_row+1
target_sheet.cell(1,2).value = 'Ration Card'
target_sheet.cell(1,3).value = 'Phone Number'


# Iterating and Searching for aadhar values in source
for row in range(2, length):
    cell = target_sheet.cell(row, 1)
    aadhar = cell.value
    data = df[['Ration-card', 'Phone-NUmber']].where(df['Aadhar-No'] == aadhar).dropna()
    ration = data['Ration-card'].values[0]
    phone = data['Phone-NUmber'].values[0]
    cell1 = target_sheet.cell(row, 2)
    cell1.value = ration
    cell2 = target_sheet.cell(row, 3)
    cell2.value = phone
target.save('completed.xlsx')
